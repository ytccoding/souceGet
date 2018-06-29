# -*- coding: utf-8 -*-
from selenium import webdriver
from time import sleep
from openpyxl import workbook ,load_workbook ,Workbook
import os ,time ,ytFuntion
from bs4 import BeautifulSoup

def sheet_date():
    sheet["F" + str(i)].value = time.strftime("%y_%m_%d") #檢查日期
    sheet["G" + str(i)].value = time.strftime("%H_%M_%S") #檢查時間

testdayFile = time.strftime("%y_%m_%d")
testdayTime  = time.strftime("%y_%m_%d_%H_%M_%S")

wb_url = load_workbook("D:\\Test\\取得原始碼\\對照表.xlsx")
sheet_url = wb_url["工作表1"] # 獲取一張表
wb = load_workbook("D:\\Test\\取得原始碼\\前台.xlsx")

if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)
os.chdir(testdayFile)
wb.save(os.getcwd() + "\\" + "前台.xlsx") #EXCEL搬家
chrome_path = "D:\selenium_driver_chrome\chromedriver.exe" #webdriver放置資料夾
webDriver = webdriver.Chrome(chrome_path)
test_web = ytFuntion.test_web(webDriver)
    
for k in range(2 ,len(sheet_url["B"])+1):
    File_name = str(sheet_url["B" + str(k)].value).strip()
    if not os.path.exists(testdayFile + "_" + File_name):    #先確認資料夾是否存在
        os.makedirs(testdayFile + "_" + File_name)
    
    wb = load_workbook("前台.xlsx")
    sheet = wb["web"] # 獲取一張表

    testWebUrl = str(sheet_url["C" + str(k)].value).strip()
    sheet["A1"].value = testWebUrl
    webDriver.get(str(testWebUrl) + str(sheet["D2"].value).strip())

    for i in range(2 ,len(sheet["B"])+1):
        if i == 10:
            test_web.elementClick("亲，请登录",3)
            test_web.elementSendKeys("input[tag=帐号]" ,6 ,text = "ytau")
            test_web.elementSendKeys("input[tag=密码]" ,6 ,text = "qwe123")
            test_web.elementClick("[class='mainColorBtn submitBtnBig ClickShade']" ,6)
            sleep(10)
        testUrl = str(testWebUrl) + str(sheet["D" + str(i)].value).strip()
        sheet["D" + str(i)].value = testUrl
        webDriver.get(testUrl)
        sleep(3)
        text = open(os.getcwd() + "\\" + testdayFile + "_" + File_name + "\\" + str(i) + "_" + \
                    str(sheet["B" + str(i)].value).strip() + str(time.strftime("%y_%m_%d_%H_%M_%S")) + \
                    ".txt" ,"w+" , encoding = "utf8")
        if webDriver.current_url == str(testUrl):
            sheet["E" + str(i)].value = "有"
            text.write(BeautifulSoup(webDriver.page_source, 'html.parser').prettify())
            sheet_date()
        else:
            sheet["E" + str(i)].value = "無此url"
            sheet_date()
        text.close()
        
    wb.save(os.getcwd() + "\\" + testdayFile + "_" + File_name + "\\" + "原始碼_chrome_" + str(time.strftime("%y_%m_%d_%H_%M_%S") + ".xlsx"))
webDriver.quit()
